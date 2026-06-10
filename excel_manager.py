import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

DEFAULT_FILENAME = "Sandhata_Internship_Log.xlsx"

def apply_borders_to_range(ws, start_row, start_col, end_row, end_col):
    """
    Applies a clean, thin border to all cells in the specified range.
    For merged cells, applying borders to all individual cells ensures
    that the borders display correctly around the outer boundary.
    """
    thin_side = Side(border_style="thin", color="BFBFBF")  # Nice clean soft gray border
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border

def initialize_sheet(ws):
    """
    Initializes a new worksheet with title and column widths.
    """
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Title in C1:J1 (Columns 3 to 10)
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "SANDHATA INTERNSHIP"
    title_cell.font = Font(name="Aptos Narrow", size=11, bold=True)
    title_cell.alignment = Alignment(vertical="center", horizontal="center")
    
    # Set column widths
    ws.column_dimensions['A'].width = 12  # Week label
    ws.column_dimensions['B'].width = 10  # Day label
    # Set columns C through J to width 12 for the updates text
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 12
        
    # Apply borders to the title block
    apply_borders_to_range(ws, 1, 3, 1, 10)

def initialize_week(ws, week_num):
    """
    Initializes the grid structure for a given week.
    Each week spans 10 rows:
    Row start: 2 + 10*(week_num - 1)
    Row end:   11 + 10*(week_num - 1)
    """
    row_start = 2 + 10 * (week_num - 1)
    row_end = 11 + 10 * (week_num - 1)
    
    # 1. Column A: WEEK label (merged A2:A11)
    ws.merge_cells(start_row=row_start, start_column=1, end_row=row_end, end_column=1)
    cell_a = ws.cell(row=row_start, column=1)
    cell_a.value = f"WEEK {week_num}"
    cell_a.font = Font(name="Aptos Narrow", size=11, bold=True)
    # text_rotation=255 makes the text oriented vertically (stacked characters)
    cell_a.alignment = Alignment(text_rotation=255, vertical="center", horizontal="center")
    
    # 2. Columns B and C-J: Days 1 to 5
    for d in range(1, 6):
        d_row_start = row_start + 2 * (d - 1)
        d_row_end = d_row_start + 1
        
        # Column B: DAY label (merged B2:B3)
        ws.merge_cells(start_row=d_row_start, start_column=2, end_row=d_row_end, end_column=2)
        cell_b = ws.cell(row=d_row_start, column=2)
        cell_b.value = f"DAY {d}"
        cell_b.font = Font(name="Aptos Narrow", size=11, bold=True)
        cell_b.alignment = Alignment(vertical="center", horizontal="center")
        
        # Columns C-J: Updates area (merged C2:J3)
        ws.merge_cells(start_row=d_row_start, start_column=3, end_row=d_row_end, end_column=10)
        cell_c = ws.cell(row=d_row_start, column=3)
        cell_c.font = Font(name="Aptos Narrow", size=11)
        cell_c.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        
    # Apply borders to all cells in the newly initialized week range
    apply_borders_to_range(ws, row_start, 1, row_end, 10)

def get_workbook(filepath):
    """
    Loads workbook if it exists, otherwise creates a new one.
    """
    dir_name = os.path.dirname(filepath)
    if dir_name:
        os.makedirs(dir_name, exist_ok=True)
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        if "Sheet1" in wb.sheetnames:
            ws = wb["Sheet1"]
        else:
            ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        initialize_sheet(ws)
    return wb, ws

def find_next_empty_slot(ws):
    """
    Scans the spreadsheet to find the next empty slot.
    Returns (week_num, day_num)
    """
    w = 1
    while True:
        row_start = 2 + 10 * (w - 1)
        cell_a = ws.cell(row=row_start, column=1)
        
        # If the week is not initialized at all, this is the first slot
        if cell_a.value is None or str(cell_a.value).strip() == "":
            return w, 1
            
        # Scan days 1 to 5
        for d in range(1, 6):
            d_row_start = row_start + 2 * (d - 1)
            cell_c = ws.cell(row=d_row_start, column=3)
            # The top-left cell of the merged updates range is Column C (3)
            if cell_c.value is None or str(cell_c.value).strip() == "":
                return w, d
                
        # If all days are filled, check next week
        w += 1

def format_entry_text(input_text):
    """
    Formats the user input into numbered items on a single line.
    Example input:
      - Finished onboarding
      - Met team lead
    Output:
      1. Finished onboarding 2. Met team lead
    """
    input_text = input_text.strip()
    if not input_text:
        return ""
        
    # If the user already wrote "1. X 2. Y" format manually, keep it
    if re.search(r'^1[\s\.)\-].*\b2[\s\.)\-]', input_text, re.DOTALL):
        return input_text
        
    lines = input_text.split('\n')
    clean_lines = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Strip common bullets (-, *, •, +) and any starting digits (like 1., 2), etc.)
        cleaned = re.sub(r'^[-*•+]\s*', '', line)
        cleaned = re.sub(r'^\d+[\s\.)\-]*\s*', '', cleaned)
        cleaned = cleaned.strip()
        if cleaned:
            clean_lines.append(cleaned)
            
    if not clean_lines:
        return ""
        
    # Format as: "1. Task One 2. Task Two"
    formatted_parts = []
    for i, line in enumerate(clean_lines, 1):
        # We put a space after the dot for professional styling
        formatted_parts.append(f"{i}. {line}")
    return " ".join(formatted_parts)

def get_existing_log(filepath, week_num, day_num):
    """
    Returns the existing log text for a specific week and day, or None.
    """
    if not os.path.exists(filepath):
        return None
    wb, ws = get_workbook(filepath)
    row_start = 2 + 10 * (week_num - 1)
    d_row_start = row_start + 2 * (day_num - 1)
    cell_c = ws.cell(row=d_row_start, column=3)
    val = cell_c.value
    wb.close()
    return val

def save_log(filepath, week_num, day_num, raw_text, append=True):
    """
    Saves the text log into the spreadsheet for the specified week and day.
    If append is True and there is already text, it merges them.
    """
    wb, ws = get_workbook(filepath)
    
    # Initialize week rows if they do not exist
    row_start = 2 + 10 * (week_num - 1)
    
    # Check if we need to initialize weeks up to the requested week_num
    max_initialized_week = 0
    r = 2
    while True:
        if ws.cell(row=r, column=1).value is not None:
            max_initialized_week += 1
            r += 10
        else:
            break
            
    for w in range(max_initialized_week + 1, week_num + 1):
        initialize_week(ws, w)
        
    # Write to column C (top-left of merged C-J range)
    d_row_start = row_start + 2 * (day_num - 1)
    cell_c = ws.cell(row=d_row_start, column=3)
    existing_value = cell_c.value
    
    if existing_value and append:
        # Split existing text into individual points
        existing_points = [p.strip() for p in re.split(r'\s*\b\d+[\s\.)\-]+\s*', str(existing_value)) if p.strip()]
        
        # Split new text into individual lines
        new_points = []
        for line in raw_text.split('\n'):
            line = line.strip()
            if not line:
                continue
            cleaned = re.sub(r'^[-*•+]\s*', '', line)
            cleaned = re.sub(r'^\d+[\s\.)\-]*\s*', '', cleaned)
            cleaned = cleaned.strip()
            if cleaned:
                new_points.append(cleaned)
                
        combined_points = existing_points + new_points
        formatted_text = " ".join(f"{i}. {point}" for i, point in enumerate(combined_points, 1))
    else:
        formatted_text = format_entry_text(raw_text)
        
    cell_c.value = formatted_text
    
    # Enforce wrap text and top alignment
    cell_c.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    cell_c.font = Font(name="Aptos Narrow", size=11)
    
    wb.save(filepath)
    wb.close()
    return formatted_text

def generate_timesheet_slots(arrival_str, work_hours=8, interval_hours=2, lunch_start_str="13:00", lunch_end_str="14:00"):
    # Parse times
    from datetime import datetime
    try:
        arrival_t = datetime.strptime(arrival_str.strip(), "%H:%M")
    except ValueError:
        # Fallback if invalid format
        arrival_t = datetime.strptime("09:00", "%H:%M")
        
    lunch_start_t = datetime.strptime(lunch_start_str, "%H:%M")
    lunch_end_t = datetime.strptime(lunch_end_str, "%H:%M")
    
    # Convert to minutes from midnight
    start_min = arrival_t.hour * 60 + arrival_t.minute
    lunch_start = lunch_start_t.hour * 60 + lunch_start_t.minute
    lunch_end = lunch_end_t.hour * 60 + lunch_end_t.minute
    lunch_duration = lunch_end - lunch_start
    
    total_work_min = work_hours * 60
    interval_min = interval_hours * 60
    
    curr = start_min
    allocated_work = 0
    slots = []
    
    while allocated_work < total_work_min:
        remaining_block_work = min(interval_min, total_work_min - allocated_work)
        
        if curr < lunch_start:
            work_before_lunch = lunch_start - curr
            if work_before_lunch >= remaining_block_work:
                slots.append({
                    "start": curr,
                    "end": curr + remaining_block_work,
                    "type": "Work",
                    "duration": remaining_block_work / 60.0
                })
                allocated_work += remaining_block_work
                curr += remaining_block_work
            else:
                if work_before_lunch > 0:
                    slots.append({
                        "start": curr,
                        "end": lunch_start,
                        "type": "Work",
                        "duration": work_before_lunch / 60.0
                    })
                    allocated_work += work_before_lunch
                
                slots.append({
                    "start": lunch_start,
                    "end": lunch_end,
                    "type": "Lunch Break",
                    "duration": lunch_duration / 60.0
                })
                
                rem = remaining_block_work - work_before_lunch
                slots.append({
                    "start": lunch_end,
                    "end": lunch_end + rem,
                    "type": "Work",
                    "duration": rem / 60.0
                })
                allocated_work += rem
                curr = lunch_end + rem
        elif curr >= lunch_start and curr < lunch_end:
            lunch_left = lunch_end - curr
            slots.append({
                "start": curr,
                "end": lunch_end,
                "type": "Lunch Break",
                "duration": lunch_left / 60.0
            })
            curr = lunch_end
        else:
            slots.append({
                "start": curr,
                "end": curr + remaining_block_work,
                "type": "Work",
                "duration": remaining_block_work / 60.0
            })
            allocated_work += remaining_block_work
            curr += remaining_block_work
            
    formatted_slots = []
    for slot in slots:
        sh = int(slot["start"] // 60)
        sm = int(slot["start"] % 60)
        eh = int(slot["end"] // 60)
        em = int(slot["end"] % 60)
        formatted_slots.append({
            "start": f"{sh:02d}:{sm:02d}",
            "end": f"{eh:02d}:{em:02d}",
            "type": slot["type"],
            "duration": slot["duration"]
        })
        
    return formatted_slots

def find_day_row_range(ws, week_num, day_num):
    for r in range(3, ws.max_row + 1):
        w_val = ws.cell(row=r, column=1).value
        d_val = ws.cell(row=r, column=2).value
        
        if w_val == f"WEEK {week_num}" and d_val == f"DAY {day_num}":
            start_row = r
            end_row = r
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row == r and merged_range.min_col == 1:
                    end_row = merged_range.max_row
                    break
            return start_row, end_row
            
    return None, None

def get_next_available_row(ws):
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, 9)):
            return r + 1
    return 3 # row 1 title, row 2 header

def initialize_timesheet_sheet(ws):
    ws.views.sheetView[0].showGridLines = True
    
    # Title in A1:H1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "SANDHATA INTERNSHIP - TIMESHEET LOG"
    title_cell.font = Font(name="Aptos Narrow", size=12, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    title_cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 26
    
    # Headers in A2:H2
    headers = ["Week", "Day", "Date", "Start Time", "End Time", "Duration (Hrs)", "Category", "Activity Log"]
    header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    header_font = Font(name="Aptos Narrow", size=11, bold=True, color="000000")
    header_align = Alignment(vertical="center", horizontal="center")
    
    ws.row_dimensions[2].height = 22
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        
    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 60
    
    apply_borders_to_range(ws, 1, 1, 2, 8)

def get_timesheet_sheet(filepath):
    dir_name = os.path.dirname(filepath)
    if dir_name:
        os.makedirs(dir_name, exist_ok=True)
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
    else:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        initialize_sheet(ws1)
        
    if "Timesheet" not in wb.sheetnames:
        ws = wb.create_sheet(title="Timesheet")
        initialize_timesheet_sheet(ws)
    else:
        ws = wb["Timesheet"]
        
    wb.active = ws
    return wb, ws

def get_or_create_day_slots(filepath, week_num, day_num, date_str, arrival_time):
    wb, ws = get_timesheet_sheet(filepath)
    
    start_row, end_row = find_day_row_range(ws, week_num, day_num)
    
    if start_row is not None:
        slots = []
        for r in range(start_row, end_row + 1):
            slots.append({
                "row": r,
                "start": ws.cell(row=r, column=4).value,
                "end": ws.cell(row=r, column=5).value,
                "duration": ws.cell(row=r, column=6).value,
                "type": ws.cell(row=r, column=7).value,
                "activity": ws.cell(row=r, column=8).value
            })
        wb.close()
        return slots
    
    raw_slots = generate_timesheet_slots(arrival_time)
    start_row = get_next_available_row(ws)
    end_row = start_row + len(raw_slots) - 1
    
    ws.cell(row=start_row, column=1, value=f"WEEK {week_num}")
    ws.cell(row=start_row, column=2, value=f"DAY {day_num}")
    ws.cell(row=start_row, column=3, value=date_str)
    
    for i, slot in enumerate(raw_slots):
        r = start_row + i
        ws.cell(row=r, column=4, value=slot["start"])
        ws.cell(row=r, column=5, value=slot["end"])
        ws.cell(row=r, column=6, value=slot["duration"])
        ws.cell(row=r, column=7, value=slot["type"])
        
        if slot["type"] == "Lunch Break":
            ws.cell(row=r, column=8, value="Lunch Break")
        else:
            ws.cell(row=r, column=8, value="")
            
    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
    ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
    
    thin_side = Side(border_style="thin", color="BFBFBF")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    lunch_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    for r in range(start_row, end_row + 1):
        is_lunch = (ws.cell(row=r, column=7).value == "Lunch Break")
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = Font(name="Aptos Narrow", size=11, bold=(c in (1, 2)))
            
            if c in (1, 2, 3, 4, 5, 6, 7):
                cell.alignment = Alignment(vertical="center", horizontal="center")
            else:
                cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
                
            if is_lunch:
                cell.fill = lunch_fill
                
    wb.save(filepath)
    wb.close()
    
    return get_or_create_day_slots(filepath, week_num, day_num, date_str, arrival_time)

def recreate_day_slots(filepath, week_num, day_num, date_str, arrival_time):
    wb, ws = get_timesheet_sheet(filepath)
    start_row, end_row = find_day_row_range(ws, week_num, day_num)
    
    if start_row is not None:
        try:
            ws.unmerge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
            ws.unmerge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
            ws.unmerge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
        except Exception:
            pass
            
        ws.delete_rows(start_row, amount=(end_row - start_row + 1))
        
    wb.save(filepath)
    wb.close()
    
    return get_or_create_day_slots(filepath, week_num, day_num, date_str, arrival_time)

def save_timesheet_slot_activity(filepath, row_idx, activity_text):
    wb = load_workbook(filepath)
    if "Timesheet" not in wb.sheetnames:
        wb.close()
        raise ValueError("Timesheet sheet does not exist.")
    ws = wb["Timesheet"]
    
    # Save the activity in column H (8)
    cell = ws.cell(row=row_idx, column=8)
    cell.value = activity_text
    cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    cell.font = Font(name="Aptos Narrow", size=11)
    
    wb.active = ws
    wb.save(filepath)
    wb.close()

def sync_timesheet_to_daily_log(filepath, week_num, day_num):
    wb = load_workbook(filepath)
    if "Timesheet" not in wb.sheetnames:
        wb.close()
        return
    ws_ts = wb["Timesheet"]
    
    start_row, end_row = find_day_row_range(ws_ts, week_num, day_num)
    if not start_row:
        wb.close()
        return
        
    activities = []
    for r in range(start_row, end_row + 1):
        cat = ws_ts.cell(row=r, column=7).value
        act = ws_ts.cell(row=r, column=8).value
        if cat == "Work" and act:
            act_str = str(act).strip()
            if act_str:
                activities.append(act_str)
                
    wb.close()
    
    if activities:
        formatted_text = ""
        for i, act in enumerate(activities, 1):
            cleaned = re.sub(r'^[-*•+]\s*', '', act)
            cleaned = re.sub(r'^\d+[\s\.)\-]*\s*', '', cleaned)
            cleaned = cleaned.strip()
            if cleaned:
                if formatted_text:
                    formatted_text += f" {i}. {cleaned}"
                else:
                    formatted_text = f"{i}. {cleaned}"
        
        save_log(filepath, week_num, day_num, formatted_text, append=False)

