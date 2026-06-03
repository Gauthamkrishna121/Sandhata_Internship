import os
import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

DEFAULT_FILENAME = "Sandhata_Internship_Log.xlsx"

def apply_borders_to_range(ws, start_row, start_col, end_row, end_col):
    """
    Applies a clean, thin border to all cells in the specified range.
    For merged cells, applying borders to all individual cells ensures
    that the borders display correctly around the outer boundary.
    """
    thin_side = Side(border_style="thin", color="BFBFBF")  # Nice clean soft gray border
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border

def initialize_sheet(ws):
    """
    Initializes a new worksheet with title and column widths.
    """
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Title in C1:J1 (Columns 3 to 10)
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=3)
    title_cell.value = "SANDHATA INTERNSHIP"
    title_cell.font = Font(name="Aptos Narrow", size=11, bold=True)
    title_cell.alignment = Alignment(vertical="center", horizontal="center")
    
    # Set column widths
    ws.column_dimensions['A'].width = 12  # Week label
    ws.column_dimensions['B'].width = 10  # Day label
    # Set columns C through J to width 12 for the updates text
    for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 12
        
    # Apply borders to the title block
    apply_borders_to_range(ws, 1, 3, 1, 10)

def initialize_week(ws, week_num):
    """
    Initializes the grid structure for a given week.
    Each week spans 10 rows:
    Row start: 2 + 10*(week_num - 1)
    Row end:   11 + 10*(week_num - 1)
    """
    row_start = 2 + 10 * (week_num - 1)
    row_end = 11 + 10 * (week_num - 1)
    
    # 1. Column A: WEEK label (merged A2:A11)
    ws.merge_cells(start_row=row_start, start_column=1, end_row=row_end, end_column=1)
    cell_a = ws.cell(row=row_start, column=1)
    cell_a.value = f"WEEK {week_num}"
    cell_a.font = Font(name="Aptos Narrow", size=11, bold=True)
    # text_rotation=255 makes the text oriented vertically (stacked characters)
    cell_a.alignment = Alignment(text_rotation=255, vertical="center", horizontal="center")
    
    # 2. Columns B and C-J: Days 1 to 5
    for d in range(1, 6):
        d_row_start = row_start + 2 * (d - 1)
        d_row_end = d_row_start + 1
        
        # Column B: DAY label (merged B2:B3)
        ws.merge_cells(start_row=d_row_start, start_column=2, end_row=d_row_end, end_column=2)
        cell_b = ws.cell(row=d_row_start, column=2)
        cell_b.value = f"DAY {d}"
        cell_b.font = Font(name="Aptos Narrow", size=11, bold=True)
        cell_b.alignment = Alignment(vertical="center", horizontal="center")
        
        # Columns C-J: Updates area (merged C2:J3)
        ws.merge_cells(start_row=d_row_start, start_column=3, end_row=d_row_end, end_column=10)
        cell_c = ws.cell(row=d_row_start, column=3)
        cell_c.font = Font(name="Aptos Narrow", size=11)
        cell_c.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        
    # Apply borders to all cells in the newly initialized week range
    apply_borders_to_range(ws, row_start, 1, row_end, 10)

def get_workbook(filepath):
    """
    Loads workbook if it exists, otherwise creates a new one.
    """
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        initialize_sheet(ws)
    return wb, ws

def find_next_empty_slot(ws):
    """
    Scans the spreadsheet to find the next empty slot.
    Returns (week_num, day_num)
    """
    w = 1
    while True:
        row_start = 2 + 10 * (w - 1)
        cell_a = ws.cell(row=row_start, column=1)
        
        # If the week is not initialized at all, this is the first slot
        if cell_a.value is None or str(cell_a.value).strip() == "":
            return w, 1
            
        # Scan days 1 to 5
        for d in range(1, 6):
            d_row_start = row_start + 2 * (d - 1)
            cell_c = ws.cell(row=d_row_start, column=3)
            # The top-left cell of the merged updates range is Column C (3)
            if cell_c.value is None or str(cell_c.value).strip() == "":
                return w, d
                
        # If all days are filled, check next week
        w += 1

def format_entry_text(input_text):
    """
    Formats the user input into numbered items on a single line.
    Example input:
      - Finished onboarding
      - Met team lead
    Output:
      1. Finished onboarding 2. Met team lead
    """
    input_text = input_text.strip()
    if not input_text:
        return ""
        
    # If the user already wrote "1. X 2. Y" format manually, keep it
    if re.search(r'^1[\s\.)\-].*\b2[\s\.)\-]', input_text, re.DOTALL):
        return input_text
        
    lines = input_text.split('\n')
    clean_lines = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Strip common bullets (-, *, •, +) and any starting digits (like 1., 2), etc.)
        cleaned = re.sub(r'^[-*•+]\s*', '', line)
        cleaned = re.sub(r'^\d+[\s\.)\-]*\s*', '', cleaned)
        cleaned = cleaned.strip()
        if cleaned:
            clean_lines.append(cleaned)
            
    if not clean_lines:
        return ""
        
    # Format as: "1. Task One 2. Task Two"
    formatted_parts = []
    for i, line in enumerate(clean_lines, 1):
        # We put a space after the dot for professional styling
        formatted_parts.append(f"{i}. {line}")
    return " ".join(formatted_parts)

def get_existing_log(filepath, week_num, day_num):
    """
    Returns the existing log text for a specific week and day, or None.
    """
    if not os.path.exists(filepath):
        return None
    wb, ws = get_workbook(filepath)
    row_start = 2 + 10 * (week_num - 1)
    d_row_start = row_start + 2 * (day_num - 1)
    cell_c = ws.cell(row=d_row_start, column=3)
    val = cell_c.value
    wb.close()
    return val

def save_log(filepath, week_num, day_num, raw_text, append=True):
    """
    Saves the text log into the spreadsheet for the specified week and day.
    If append is True and there is already text, it merges them.
    """
    wb, ws = get_workbook(filepath)
    
    # Initialize week rows if they do not exist
    row_start = 2 + 10 * (week_num - 1)
    
    # Check if we need to initialize weeks up to the requested week_num
    max_initialized_week = 0
    r = 2
    while True:
        if ws.cell(row=r, column=1).value is not None:
            max_initialized_week += 1
            r += 10
        else:
            break
            
    for w in range(max_initialized_week + 1, week_num + 1):
        initialize_week(ws, w)
        
    # Write to column C (top-left of merged C-J range)
    d_row_start = row_start + 2 * (day_num - 1)
    cell_c = ws.cell(row=d_row_start, column=3)
    existing_value = cell_c.value
    
    if existing_value and append:
        # Split existing text into individual points
        existing_points = [p.strip() for p in re.split(r'\s*\b\d+[\s\.)\-]+\s*', str(existing_value)) if p.strip()]
        
        # Split new text into individual lines
        new_points = []
        for line in raw_text.split('\n'):
            line = line.strip()
            if not line:
                continue
            cleaned = re.sub(r'^[-*•+]\s*', '', line)
            cleaned = re.sub(r'^\d+[\s\.)\-]*\s*', '', cleaned)
            cleaned = cleaned.strip()
            if cleaned:
                new_points.append(cleaned)
                
        combined_points = existing_points + new_points
        formatted_text = " ".join(f"{i}. {point}" for i, point in enumerate(combined_points, 1))
    else:
        formatted_text = format_entry_text(raw_text)
        
    cell_c.value = formatted_text
    
    # Enforce wrap text and top alignment
    cell_c.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    cell_c.font = Font(name="Aptos Narrow", size=11)
    
    wb.save(filepath)
    wb.close()
    return formatted_text
